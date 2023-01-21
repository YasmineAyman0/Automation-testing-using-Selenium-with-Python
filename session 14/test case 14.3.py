# test case 14.3 excel utils file
import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(excel_file, sheetname):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheetname]
    return (sheet.max_rows)


def getColumnCount(excel_file, sheetname):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheetname]
    return (sheet.max_column)


def readData(excel_file, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheetname]
    return sheet.cell(rownum, colnum).value


def writeData(excel_file, sheetname, rownum, colnum, data):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheetname]
    sheet.cell(rownum, colnum).value = data
    workbook.save(excel_file)


def fillGreenColor(excel_file, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheetname]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rownum, colnum).fill = greenFill
    workbook.save(excel_file)


def fillRedColor(excel_file, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheetname]
    redFill = PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(rownum, colnum).fill = redFill
    workbook.save(excel_file)
