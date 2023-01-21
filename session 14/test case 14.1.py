# test case 14.1 Reading data from excel
import openpyxl

excel_file = "E:\Interviews\Automation\python\python testing files\session 14\data.xlsx"
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook["Sheet1"]
rows = sheet.max_row
cols = sheet.max_column
# Reading all the the rows & columns from excel sheet
for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(r, c).value, end='                    ')
        print()
